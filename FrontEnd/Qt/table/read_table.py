from PyQt6.QtWidgets import (QApplication,QWidget,QMainWindow,
                             QTableWidget,QVBoxLayout,
                             QPushButton,QTableWidgetItem)
from PyQt6.QtCore import Qt,QUrl
from PyQt6.QtGui import QDesktopServices,QIcon
import pandas as pd

class MyApp(QWidget):
    def __init__(self):
        super().__init__()
        self.resize(700,500)
        layout=QVBoxLayout()
        self.setLayout(layout)
        self.table=QTableWidget()
        layout.addWidget(self.table)
        self.button=QPushButton('Load table')
        layout.addWidget(self.button)
        self.button.clicked.connect(self.print_data)
    def import_excel_table(self):
        from openpyxl import load_workbook
        file_path='table\multilang.xlsx'
        try:
            workbook=load_workbook(file_path)
        except FileNotFoundError:
            return f"Error: File not found at {file_path}"
        except Exception as e:
            return f"An error occurred: {e}"
        sheet=workbook.active
        headers=[cell.value for cell in sheet[1]]
        dicts=[]
        for row in sheet.iter_rows(min_row=2,values_only=True):
            row_dict=dict(zip(headers,row))
            dicts.append(row_dict)
        workbook.close()
        phrases=[x['Front'] for x in dicts if 'Front' in x]
        words=[x['Back'] for x in dicts if 'Back' in x]
        trans_phrases=[x['trans_phrases'] for x in dicts 
                            if 'trans_phrases' in x]
        trans_words=[x['trans_words'] for x in dicts
                            if 'trans_words' in x]
        langs=[x['Lang'] for x in dicts if 'Lang' in x]
        return phrases,words,trans_phrases,trans_words,langs
    def OpenChatGPT(self):
        QDesktopServices.openUrl(QUrl('https://chat.openai.com/'))
    def print_data(self):
        phrases,words,trans_phrases,trans_words,langs=self.import_excel_table()
        self.table.setRowCount(len(phrases))
        self.table.setColumnCount(4)
        self.table.setHorizontalHeaderLabels(['Phrases','Words','Langs','ChatGPT'])
        self.table.cellDoubleClicked.connect(self.OpenChatGPT)
        for index,phrase in enumerate(phrases):
            item=QTableWidgetItem(phrase)
            self.table.setItem(index,0,item)
        for index,word in enumerate(words):
            item=QTableWidgetItem(word)
            self.table.setItem(index,1,item)
        for index,lang in enumerate(langs):
            item=QTableWidgetItem(lang)
            self.table.setItem(index,2,item)
        for index,_ in enumerate(phrases):
            item=QTableWidgetItem('')
            item.setIcon(QIcon('table\chatgpt_logo.png'))
            self.table.setItem(index,3,item)
        #
app=QApplication([])
myapp=MyApp()
myapp.show()
app.exec()